"""
Ekstraktor seed Dasbort ESG.

Membaca sheet "Index ESG" dari workbook master dan menulis seed-esg.json:
- dimensi 13 metrik (bobot, target, pilar, kategori, arah, tipe rumus)
- nilai bulanan: aktual (Blok A), %achievement (Blok B), point (Blok C)
- expected sum (baris 51) & total/YTD POINT (baris 52) untuk uji paritas

Jalankan:  py src/dasbort-esg/scripts/extract_seed.py
Sumber Excel bersifat privat & TIDAK di-deploy; hasilnya (seed-esg.json) yang dipakai app.
"""
import json
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)                       # src/dasbort-esg
SRC_XLSX = os.path.abspath(os.path.join(
    ROOT, "..", "..", "ESG", "Environment & ESG Report 2026_PELH Juni.xlsx"))
OUT = os.path.join(ROOT, "data", "seed-esg.json")

MONTHS = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun",
          "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]

# id, nama, kategori headline, pilar, satuan, arah (lower=makin kecil makin baik), tipe rumus Layer-1
# Baris di sheet Index ESG: Blok A = 4+i, Blok B = 21+i, Blok C = 38+i
METRICS = [
    ("ghg_int",        "GHG Intensity Reduction",             "Energi & Emisi",  "E", "%",          "higher", "ratio"),
    ("fuel_ratio",     "Fuel Ratio",                          "Energi & Emisi",  "E", "liter/bcm",  "lower",  "fuel_inverse"),
    ("renewable",      "Renewable Energy Mix",                "Energi Terbarukan","E", "%",          "higher", "ratio"),
    ("water_int",      "Water Withdrawal Intensity Reduction","Air",             "E", "%",          "higher", "ratio"),
    ("air_bersih",     "Pemakaian Air Bersih",                "Air",             "E", "%",          "lower",  "direct"),
    ("air_konservasi", "Pemanfaatan Air Konservasi",          "Air",             "E", "%",          "higher", "direct"),
    ("flowmeter",      "Instalasi Flowmeter",                 "Air",             "E", "%",          "higher", "direct"),
    ("waste_diverted", "Solid Waste Diverted",                "Limbah",          "E", "%",          "higher", "ratio"),
    ("limbah_b3",      "Limbah B3 Terolah",                   "Limbah",          "E", "%",          "higher", "direct"),
    ("limbah_domestik","Limbah Domestik Terolah",             "Limbah",          "E", "%",          "higher", "direct"),
    ("limbah_produksi","Limbah Produksi Terolah",             "Limbah",          "E", "%",          "higher", "direct"),
    ("ltifr",          "LTI-FR",                              "K3 (LTI-FR)",     "S", "unitless",   "lower",  "direct"),
    ("comdev",         "Community Development Beneficiaries",  "Community Development","S","orang",   "higher", "ratio_cumulative"),
]

COLS = ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]


def num(v):
    """Kembalikan float bila numerik, else None (mis. '#DIV/0!')."""
    return float(v) if isinstance(v, (int, float)) else None


def main():
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb["Index ESG"]

    metrics_out = []
    data_out = {}
    for i, (mid, name, kat, pilar, satuan, arah, rumus) in enumerate(METRICS):
        rowA, rowB, rowC = 4 + i, 21 + i, 38 + i
        weight = num(ws[f"C{rowC}"].value) or 0.0
        target = num(ws[f"D{rowA}"].value)
        aktual = [num(ws[f"{c}{rowA}"].value) for c in COLS]
        achiev = [num(ws[f"{c}{rowB}"].value) for c in COLS]
        point = [num(ws[f"{c}{rowC}"].value) for c in COLS]
        metrics_out.append({
            "id": mid, "no": i + 1, "name": name, "kategori": kat,
            "pilar": pilar, "satuan": satuan, "arah": arah,
            "bobot": round(weight, 6),
            "target2026": target, "rumus": rumus, "aktif": True,
        })
        data_out[mid] = {"aktual": aktual, "achievement": achiev, "point": point}

    expected = {
        "sum": [num(ws[f"{c}51"].value) for c in COLS],
        "total": [num(ws[f"{c}52"].value) for c in COLS],
    }

    seed = {
        "meta": {
            "sumber": "Environment & ESG Report 2026_PELH Juni.xlsx / sheet 'Index ESG'",
            "dataAsOf": "2026-06",
            "target": 1.0,
            "cap": 1.25,
            "tahun": 2026,
        },
        "months": MONTHS,
        "metrics": metrics_out,
        "data": data_out,
        "expected": expected,
    }

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(seed, f, ensure_ascii=False, indent=2)

    # Seed untuk Apps Script (agar Sheet bisa dibangun+diisi di akun mana pun,
    # tanpa duplikasi manual — sumber tunggal tetap Excel).
    gs_dir = os.path.join(ROOT, "apps-script")
    os.makedirs(gs_dir, exist_ok=True)
    gs_path = os.path.join(gs_dir, "seed_data.gs")
    with open(gs_path, "w", encoding="utf-8") as f:
        f.write("// GENERATED oleh scripts/extract_seed.py — JANGAN edit tangan.\n")
        f.write("// Seed metrik + data Jan-Jun 2026 untuk setupAwal().\n")
        f.write("var SEED_ESG = " + json.dumps(seed, ensure_ascii=False) + ";\n")
    print(f"OK -> {gs_path}")

    bobot_total = round(sum(m["bobot"] for m in metrics_out), 4)
    print(f"OK -> {OUT}")
    print(f"metrik: {len(metrics_out)}  total bobot: {bobot_total}")
    print("expected total (Jan-Jun):",
          [round(x, 4) if x is not None else None for x in expected["total"][:6]])


if __name__ == "__main__":
    main()
